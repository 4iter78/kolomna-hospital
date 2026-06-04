from datetime import datetime

from flask import (
    Blueprint,
    render_template,
    request,
    session
)

from app import db_connection
from decorators import access_control
from models.Departments import Departments
from models.MaterialBalances import MaterialBalances
from models.MaterialOperations import MaterialOperations
from models.MaterialUnits import MaterialUnits
from models.MedicalMaterials import MedicalMaterials
from models.Users import Users
from io import BytesIO
from flask import send_file
from openpyxl import Workbook

db = db_connection

material_balances_controller = Blueprint(
    'material_balances_controller',
    __name__
)


def autosize_columns(ws):
    for column in ws.columns:
        max_length = max(
            len(str(cell.value))
            for cell in column
            if cell.value
        )
        ws.column_dimensions[
            column[0].column_letter
        ].width = min(max_length + 3, 50)


@material_balances_controller.route('/storage',methods=['GET'])
@access_control('storage')
def handle_storage():

    balances = MaterialBalances.query.all()
    available = []
    for balance in balances:
        material = MedicalMaterials.query.get(
            balance.medical_material_id
        )
        material_unit = None
        if material:
            material_unit = MaterialUnits.query.get(
                material.material_unit_id
            )
        department = Departments.query.get(
            balance.department_id
        )
        txt_balance = {
            "id": balance.id,
            "medical_material_id":
                balance.medical_material_id,
            "medical_material":
                material.name if material else '',
            "current_quantity":
                balance.current_quantity,
            "table_quantity":
                f'{balance.current_quantity} {material_unit.short_name if material else ''}',
            "document_number":
                '',
            "department_id":
                balance.department_id,
            "department":
                department.name if department else '',
            "operation_date":
                datetime.now().strftime("%Y-%m-%dT%H:%M"),
            "last_updated":
                balance.last_updated,
            "description":
                '',
            "is_not_storage":
                department.name != 'Склад'
        }
        available.append(txt_balance)
    materials = [
        {
            "id": material.id,
            "name": material.name
        }
        for material in MedicalMaterials.query.all()
    ]
    departments = [
        {
            "id": department.id,
            "name": department.name
        }
        for department in Departments.query.all()
    ]
    return render_template(
        'storage.html',
        title='Медицинские материалы в наличии (Хранение)',
        available=available,
        materials=materials,
        departments=departments,
        count=len(available)
    )


@material_balances_controller.route('/storage/export')
@access_control('storage')
def export_storage_excel():

    balances = MaterialBalances.query.all()

    rows = []

    for balance in balances:
        material = MedicalMaterials.query.get(
            balance.medical_material_id
        )
        department = Departments.query.get(
            balance.department_id
        )
        unit = None
        if material:
            unit = MaterialUnits.query.get(
                material.material_unit_id
            )
        rows.append({
            'department':
                department.name if department else '',
            'material':
                material.name if material else '',
            'quantity':
                f'{balance.current_quantity} {unit.short_name if unit else ""}',
            'last_updated':
                (
                    balance.last_updated.strftime(
                        '%d.%m.%Y %H:%M'
                    )
                    if balance.last_updated
                    else ''
                )
        })

    # Сортировка:
    # Отделение -> Материал

    rows.sort(
        key=lambda x: (
            x['department'].lower(),
            x['material'].lower()
        )
    )

    wb = Workbook()
    ws = wb.active
    ws.title = f"{datetime.now().strftime("%Y-%m-%d")}"

    ws.append([
        "№ п/п",
        "Отделение",
        "Медицинский материал",
        "Количество",
        "Дата обновления"
    ])

    for index, row in enumerate(rows,start=1):

        ws.append([
            index,
            row['department'],
            row['material'],
            row['quantity'],
            row['last_updated']
        ])

    autosize_columns(ws)
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    return send_file(
        output,
        as_attachment=True,
        download_name="material_balances.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )


@material_balances_controller.route('/issued',methods=['GET'])
@access_control('issued')
def handle_issued():

    issued_operations = MaterialOperations.query.filter_by(
        is_issued=True
    ).all()
    issued = []
    for operation in issued_operations:
        material = MedicalMaterials.query.get(
            operation.medical_material_id
        )
        current_user = Users.query.get(
            operation.current_user_id
        )
        material_unit = None
        if material:
            material_unit = MaterialUnits.query.get(
                material.material_unit_id
            )
        department = Departments.query.get(
            operation.department_id
        )
        txt_operation = {
            "id": operation.id,
            "department":
                department.name if department else '',
            "current_user":
                f'{current_user.surname} {current_user.name} {current_user.second_name}',
            "medical_material":
                material.name if material else '',
            "quantity": operation.quantity,
            "table_quantity": f'{operation.quantity} {material_unit.short_name if material else ''}',
            "document_number":
                operation.document_number,
            "operation_date":
                operation.operation_date
        }
        issued.append(txt_operation)
    materials = [
        {
            "id": material.id,
            "name": material.name
        }
        for material in MedicalMaterials.query.all()
    ]
    departments = [
        {
            "id": department.id,
            "name": department.name
        }
        for department in Departments.query.all()
    ]
    return render_template(
        'issued.html',
        title='Выданные медицинские материалы',
        issued=issued,
        materials=materials,
        departments=departments,
        count=len(issued)
    )

@material_balances_controller.route('/issued/<issued_id>',methods=['PUT'])
@access_control('issued')
def handle_issued_item(issued_id):

    balance = MaterialBalances.query.get_or_404(
        issued_id
    )

    material = MedicalMaterials.query.get_or_404(
        balance.medical_material_id
    )

    if request.method == 'PUT':
        try:
            data = request.get_json()
            new_material_operation = MaterialOperations(
                medical_material_id=balance.medical_material_id,
                quantity=data['quantity'],
                document_number=data['document_number'],
                department_id=balance.department_id,
                current_user_id=session.get('user_id'),
                operation_date=data['operation_date'],
                description=data['description'],
                is_issued=True
            )
            db.session.add(new_material_operation)
            db.session.commit()
            return {"success": True, "message": f"Материал {material.name} успешно выдан."}
        except Exception as e:
            db.session.rollback()
            print(e)
            return {"success": False, "message": f"Материал {material.name} не может быть выдан. {str(e)}"}, 400


@material_balances_controller.route('/issued/export')
@access_control('issued')
def export_issued_excel():

    issued_operations = (
        MaterialOperations.query
        .filter_by(is_issued=True)
        .all()
    )

    rows = []

    for operation in issued_operations:

        material = MedicalMaterials.query.get(
            operation.medical_material_id
        )
        current_user = Users.query.get(
            operation.current_user_id
        )
        department = Departments.query.get(
            operation.department_id
        )
        material_unit = None

        if material:
            material_unit = MaterialUnits.query.get(
                material.material_unit_id
            )

        rows.append({
            'department':
                department.name if department else '',
            'user':
                (
                    f'{current_user.surname} '
                    f'{current_user.name} '
                    f'{current_user.second_name}'
                ) if current_user else '',
            'material':
                material.name if material else '',
            'quantity':
                (
                    f'{operation.quantity} '
                    f'{material_unit.short_name if material_unit else ""}'
                ),
            'document_number':
                operation.document_number or '',
            'operation_date':
                (
                    operation.operation_date.strftime(
                        '%d.%m.%Y %H:%M'
                    )
                    if operation.operation_date
                    else ''
                )
        })

    # Сортировка:
    # Отделение -> Материал
    rows.sort(
        key=lambda x: (
            x['department'].lower(),
            x['material'].lower()
        )
    )

    wb = Workbook()
    ws = wb.active
    ws.title = datetime.now().strftime(
        "%Y-%m-%d"
    )

    ws.append([
        "№ п/п",
        "Отделение",
        "Выдал",
        "Материал",
        "Количество",
        "№ документа",
        "Дата операции"
    ])

    for index, row in enumerate(
        rows,
        start=1
    ):

        ws.append([
            index,
            row['department'],
            row['user'],
            row['material'],
            row['quantity'],
            row['document_number'],
            row['operation_date']
        ])

    autosize_columns(ws)
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    return send_file(
        output,
        as_attachment=True,
        download_name='issued_materials.xlsx',
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )


@material_balances_controller.route('/written_off/<written_off_id>',methods=['PUT'])
@access_control('written_off')
def handle_written_off_item(written_off_id):

    balance = MaterialBalances.query.get_or_404(
        written_off_id
    )

    material = MedicalMaterials.query.get_or_404(
        balance.medical_material_id
    )

    if request.method == 'PUT':
        try:
            data = request.get_json()
            new_material_operation = MaterialOperations(
                medical_material_id=balance.medical_material_id,
                quantity=data['quantity'],
                document_number=data['document_number'],
                department_id=balance.department_id,
                current_user_id=session.get('user_id'),
                operation_date=data['operation_date'],
                description=data['description'],
                is_written_off=True
            )
            db.session.add(new_material_operation)
            db.session.commit()
            return {"success": True, "message": f"Материал #{material.name} успешно списан."}
        except Exception as e:
            db.session.rollback()
            print(e)
            return {"success": False, "message": f"Материал {material.name} не может быть списан. {str(e)}"}, 400


@material_balances_controller.route('/written_off',methods=['GET'])
@access_control('written_off')
def handle_written_off():

    written_off_operations = MaterialOperations.query.filter_by(
        is_written_off=True
    ).all()

    written_off = []
    for operation in written_off_operations:
        material = MedicalMaterials.query.get(
            operation.medical_material_id
        )
        current_user = Users.query.get(
            operation.current_user_id
        )
        material_unit = None
        if material:
            material_unit = MaterialUnits.query.get(
                material.material_unit_id
            )
        department = Departments.query.get(
            operation.department_id
        )
        txt_operation = {
            "id": operation.id,
            "department":
                department.name if department else '',
            "current_user":
                f'{current_user.surname} {current_user.name} {current_user.second_name}',
            "medical_material":
                material.name if material else '',
            "quantity": operation.quantity,
            "table_quantity": f'{operation.quantity} {material_unit.short_name if material else ''}',
            "document_number":
                operation.document_number,
            "operation_date":
                operation.operation_date
        }
        written_off.append(txt_operation)
    materials = [
        {
            "id": material.id,
            "name": material.name
        }
        for material in MedicalMaterials.query.all()
    ]
    departments = [
        {
            "id": department.id,
            "name": department.name
        }
        for department in Departments.query.all()
    ]
    return render_template(
        'written_off.html',
        title='Списанные медицинские материалы',
        written_off=written_off,
        materials=materials,
        departments=departments,
        count=len(written_off)
    )

@material_balances_controller.route('/written_off/export')
@access_control('written_off')
def export_written_off_excel():
    written_off_ops = (MaterialOperations.query.filter_by(is_written_off=True).all())

    rows = []

    for op in written_off_ops:

        material = MedicalMaterials.query.get(
            op.medical_material_id
        )
        user = Users.query.get(
            op.current_user_id
        )
        dept = Departments.query.get(
            op.department_id
        )
        unit = None
        if material:
            unit = MaterialUnits.query.get(
                material.material_unit_id
            )

        rows.append({
            'department':
                dept.name if dept else '',
            'user':
                (
                    f'{user.surname} '
                    f'{user.name} '
                    f'{user.second_name}'
                ) if user else '',
            'material':
                material.name if material else '',
            'quantity':
                f'{op.quantity} '
                f'{unit.short_name if unit else ""}',
            'document_number':
                op.document_number or '',
            'operation_date':
                (op.operation_date.strftime('%d.%m.%Y %H:%M')
                    if op.operation_date
                    else ''
                )
        })

    # Сортировка:
    # сначала отделение,
    # потом материал

    rows.sort(
        key=lambda x: (
            x['department'].lower(),
            x['material'].lower()
        )
    )

    wb = Workbook()

    ws = wb.active

    ws.title = datetime.now().strftime("%Y-%m-%d")

    ws.append([
        "№ п/п",
        "Отделение",
        "Выдал",
        "Материал",
        "Количество",
        "№ документа",
        "Дата операции"
    ])

    for index, row in enumerate(rows,start=1):
        ws.append([
            index,
            row['department'],
            row['user'],
            row['material'],
            row['quantity'],
            row['document_number'],
            row['operation_date']
        ])

    autosize_columns(ws)
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    return send_file(
        output,
        as_attachment=True,
        download_name="written_off_materials.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
